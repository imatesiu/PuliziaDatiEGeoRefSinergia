#!/usr/bin/env python3
from __future__ import annotations

import csv
import json
import re
import shutil
import time
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable
from urllib.error import HTTPError, URLError
from urllib.parse import urlencode
from urllib.request import Request, urlopen

from openpyxl import load_workbook


PLACEHOLDER_VALUES = {
    "*INDIRIZZO ASSENTE",
    "....",
    "A CYBO",
    "CARRARA",
}
CF_RE = re.compile(r"^[A-Z]{6}\d{2}[A-Z]\d{2}[A-Z]\d{3}[A-Z]$")
SEARCH_PREFIX_RE = re.compile(
    r"^(?:LUOGO\s+DETTO|LOCALIT[ÀA']|LOC\.?|LDT)\s+",
    re.IGNORECASE,
)
TRAILING_CITY_RE = re.compile(
    r"\b\d+[A-Z/.\-]*\s+([A-ZÀ-ÖØ-Ý][A-ZÀ-ÖØ-Ý' -]+)$",
    re.IGNORECASE,
)
NOMINATIM_URL = "https://nominatim.openstreetmap.org/search"
DEFAULT_USER_AGENT = "PuliziaDatiSinergia-Geocoder/1.0 (+local-script)"


@dataclass
class Record:
    row_number: int
    data: dict[str, object]
    category: str
    issues: list[str]


def clean_value(value: object) -> object:
    if isinstance(value, str):
        stripped = value.strip()
        return stripped if stripped else ""
    return value


def normalized_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip().upper()


def normalize_search_text(value: object) -> str:
    if value in ("", None):
        return ""
    text = re.sub(r"\s+", " ", str(value).strip())
    while True:
        cleaned = SEARCH_PREFIX_RE.sub("", text).strip()
        if cleaned == text:
            return cleaned
        text = cleaned


def compact_spaces(value: object) -> str:
    if value in ("", None):
        return ""
    return re.sub(r"\s+", " ", str(value).strip())


def normalized_address_key(value: object) -> str:
    return compact_spaces(value).upper()


def infer_alternate_city(row: dict[str, str]) -> str:
    raw = compact_spaces(row.get("INDIRES", ""))
    primary_city = normalized_text(row.get("COMUNE_NEW", ""))
    match = TRAILING_CITY_RE.search(raw.upper())
    if not match:
        return ""

    candidate = compact_spaces(match.group(1))
    candidate_upper = normalized_text(candidate)
    if not candidate_upper or candidate_upper == primary_city:
        return ""
    if candidate_upper in {"INT", "INTERNO", "SCALA", "PIANO", "TERRA", "BIS", "TER"}:
        return ""
    if any(char.isdigit() for char in candidate_upper):
        return ""
    return candidate.title()


def normalize_headers(raw_headers: list[object]) -> list[str]:
    counts: dict[str, int] = {}
    headers: list[str] = []
    for raw_header in raw_headers:
        header = clean_value(raw_header)
        base_name = str(header) if header else "NOTE"
        counts[base_name] = counts.get(base_name, 0) + 1
        if counts[base_name] == 1:
            headers.append(base_name)
        else:
            headers.append(f"{base_name}_{counts[base_name]}")
    return headers


def load_rows_from_excel(input_file: Path, sheet_name: str | None) -> tuple[str, list[str], list[dict[str, object]]]:
    workbook = load_workbook(input_file, data_only=True)
    sheet = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]
    headers = normalize_headers(
        [sheet.cell(row=1, column=idx).value for idx in range(1, sheet.max_column + 1)]
    )

    rows: list[dict[str, object]] = []
    for row_idx in range(2, sheet.max_row + 1):
        values = [
            clean_value(sheet.cell(row=row_idx, column=col_idx).value)
            for col_idx in range(1, sheet.max_column + 1)
        ]
        if not any(value not in ("", None) for value in values):
            continue
        row = dict(zip(headers, values))
        row["_row"] = row_idx
        rows.append(row)
    return sheet.title, headers, rows


def detect_csv_dialect(csv_path: Path) -> csv.Dialect:
    with csv_path.open("r", encoding="utf-8-sig", newline="") as handle:
        sample = handle.read(4096)
        handle.seek(0)
        try:
            return csv.Sniffer().sniff(sample, delimiters=",;\t|")
        except csv.Error:
            return csv.excel


def load_rows_from_csv(input_file: Path) -> tuple[str, list[str], list[dict[str, object]]]:
    dialect = detect_csv_dialect(input_file)
    with input_file.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle, dialect=dialect)
        try:
            raw_headers = next(reader)
        except StopIteration as exc:
            raise RuntimeError(f"CSV vuoto: {input_file}") from exc

        headers = normalize_headers(raw_headers)
        rows: list[dict[str, object]] = []
        for row_idx, values in enumerate(reader, start=2):
            padded = values[: len(headers)] + [""] * max(0, len(headers) - len(values))
            cleaned = [clean_value(value) for value in padded]
            if not any(value not in ("", None) for value in cleaned):
                continue
            row = dict(zip(headers, cleaned))
            row["_row"] = row_idx
            rows.append(row)
    return "CSV", headers, rows


def load_rows(input_file: Path, sheet_name: str | None = None) -> tuple[str, list[str], list[dict[str, object]]]:
    suffix = input_file.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        return load_rows_from_excel(input_file, sheet_name)
    if suffix == ".csv":
        return load_rows_from_csv(input_file)
    raise RuntimeError(f"Formato non supportato: {input_file.suffix}")


def build_duplicate_map(rows: Iterable[dict[str, object]]) -> dict[str, list[int]]:
    duplicate_map: dict[str, list[int]] = defaultdict(list)
    for row in rows:
        address = normalized_text(row.get("INDIRIZZO NORMALIZZATO"))
        if address:
            duplicate_map[address].append(int(row["_row"]))
    return {
        address: positions
        for address, positions in duplicate_map.items()
        if len(positions) > 1
    }


def has_note(row: dict[str, object]) -> bool:
    for key, value in row.items():
        if key == "NOTE" or (isinstance(key, str) and key.startswith("NOTE_")):
            if value not in ("", None):
                return True
    return False


def get_primary_note(row: dict[str, object]) -> str:
    note = row.get("NOTE")
    if note not in ("", None):
        return str(note)
    for key, value in row.items():
        if isinstance(key, str) and key.startswith("NOTE_") and value not in ("", None):
            return str(value)
    return ""


def is_zero_civico(value: object) -> bool:
    if value == 0:
        return True
    if value in ("", None):
        return False
    return str(value).strip() in {"0", "0.0"}


def classify_row(row: dict[str, object], duplicate_map: dict[str, list[int]]) -> Record:
    issues: list[str] = []
    raw_address = normalized_text(row.get("INDIRES"))
    address = normalized_text(row.get("INDIRIZZO NORMALIZZATO"))
    row_number = int(row["_row"])

    has_structure = all(
        row.get(key) not in ("", None)
        for key in ("TYPE", "CIVICO_NORM", "NUMERO CIVICO", "COMUNE_NEW")
    )

    if raw_address in PLACEHOLDER_VALUES:
        issues.append("valore_segnaposto")
    if CF_RE.match(raw_address):
        issues.append("codice_fiscale")
    if has_note(row):
        issues.append("nota_presente")
    if is_zero_civico(row.get("NUMERO CIVICO")):
        issues.append("civico_zero")
    if address and address in duplicate_map:
        issues.append("indirizzo_duplicato")
    if not has_structure:
        issues.append("dati_strutturati_mancanti")

    if "codice_fiscale" in issues or "valore_segnaposto" in issues:
        category = "scarti"
    elif "dati_strutturati_mancanti" in issues:
        category = "scarti"
    else:
        category = "validi"

    return Record(row_number=row_number, data=row, category=category, issues=issues)


def categorize_records(records: list[Record]) -> dict[str, list[Record]]:
    return {
        "validi": [record for record in records if record.category == "validi"],
        "da_verificare": [record for record in records if record.category == "da_verificare"],
        "scarti": [record for record in records if record.category == "scarti"],
    }


def write_records_csv(path: Path, headers: list[str], records: list[Record]) -> None:
    fieldnames = ["_row", "_category", "_issues"] + headers
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for record in records:
            row = {
                "_row": record.row_number,
                "_category": record.category,
                "_issues": ", ".join(record.issues),
            }
            row.update({header: record.data.get(header, "") for header in headers})
            writer.writerow(row)


def write_summary(
    path: Path,
    input_file: Path,
    source_name: str,
    headers: list[str],
    records: list[Record],
    duplicate_map: dict[str, list[int]],
) -> None:
    counts = Counter(record.category for record in records)
    filled_columns = {
        header: sum(1 for record in records if record.data.get(header) not in ("", None))
        for header in headers
    }
    lines = [
        f"File analizzato: {input_file}",
        f"Origine: {source_name}",
        f"Righe dati: {len(records)}",
        "",
        "Colonne:",
    ]
    lines.extend(f"- {header}: {filled_columns[header]} valori non vuoti" for header in headers)
    lines.extend(
        [
            "",
            "Classificazione:",
            f"- validi: {counts.get('validi', 0)}",
            f"- da_verificare: {counts.get('da_verificare', 0)}",
            f"- scarti: {counts.get('scarti', 0)}",
            "",
            "Duplicati su INDIRIZZO NORMALIZZATO:",
        ]
    )
    if duplicate_map:
        lines.extend(
            f"- {address}: righe {positions}"
            for address, positions in sorted(duplicate_map.items())
        )
    else:
        lines.append("- nessuno")

    lines.extend(["", "Dettaglio righe da verificare/scartare:"])
    for record in records:
        if record.category == "validi":
            continue
        address = record.data.get("INDIRES", "")
        issues = ", ".join(record.issues) if record.issues else "-"
        lines.append(f"- riga {record.row_number}: {address} [{record.category}] -> {issues}")

    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_analysis_json(path: Path, records: list[Record], duplicate_map: dict[str, list[int]]) -> None:
    payload = {
        "summary": {
            "rows": len(records),
            "categories": dict(Counter(record.category for record in records)),
            "duplicates": duplicate_map,
        },
        "records": [
            {
                "row_number": record.row_number,
                "category": record.category,
                "issues": record.issues,
                "data": {key: value for key, value in record.data.items() if key != "_row"},
            }
            for record in records
        ],
    }
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def analyze_input_file(input_file: Path, output_dir: Path, sheet_name: str | None = None) -> dict[str, Any]:
    source_name, headers, rows = load_rows(input_file, sheet_name=sheet_name)
    duplicate_map = build_duplicate_map(rows)
    records = [classify_row(row, duplicate_map) for row in rows]
    categories = categorize_records(records)

    output_dir.mkdir(parents=True, exist_ok=True)
    stem = input_file.stem
    paths = {
        "summary": output_dir / f"{stem}_summary.txt",
        "analysis": output_dir / f"{stem}_analysis.json",
    }

    for category, category_records in categories.items():
        path = output_dir / f"{stem}_{category}.csv"
        paths[category] = path
        write_records_csv(path, headers, category_records)

    write_summary(
        paths["summary"],
        input_file=input_file,
        source_name=source_name,
        headers=headers,
        records=records,
        duplicate_map=duplicate_map,
    )
    write_analysis_json(paths["analysis"], records, duplicate_map)

    return {
        "input_file": input_file,
        "source_name": source_name,
        "headers": headers,
        "rows": rows,
        "records": records,
        "categories": categories,
        "duplicate_map": duplicate_map,
        "paths": paths,
    }


def load_cache(cache_path: Path) -> dict[str, Any]:
    if not cache_path.exists():
        return {}
    with cache_path.open("r", encoding="utf-8") as handle:
        return json.load(handle)


def save_cache(cache_path: Path, cache: dict[str, Any]) -> None:
    cache_path.parent.mkdir(parents=True, exist_ok=True)
    cache_path.write_text(json.dumps(cache, ensure_ascii=False, indent=2), encoding="utf-8")


def read_csv_rows(csv_path: Path) -> tuple[list[str], list[dict[str, str]]]:
    with csv_path.open("r", newline="", encoding="utf-8-sig") as handle:
        reader = csv.DictReader(handle)
        rows = list(reader)
        if reader.fieldnames is None:
            raise RuntimeError(f"CSV senza intestazioni: {csv_path}")
        return list(reader.fieldnames), rows


def build_structured_params(
    row: dict[str, str],
    *,
    city: str | None = None,
    country: str,
    country_code: str,
    email: str | None,
) -> dict[str, str]:
    street_name = normalize_search_text(row.get("CIVICO_NORM", ""))
    house_number = row.get("NUMERO CIVICO", "").strip()
    city = compact_spaces(city or row.get("COMUNE_NEW", ""))
    params = {
        "street": f"{house_number} {street_name}".strip(),
        "city": city,
        "country": country,
        "countrycodes": country_code,
        "format": "jsonv2",
        "limit": "1",
        "addressdetails": "1",
    }
    if email:
        params["email"] = email
    return params


def build_freeform_params(
    row: dict[str, str],
    *,
    city: str | None = None,
    country: str,
    country_code: str,
    email: str | None,
    road_only: bool = False,
) -> dict[str, str]:
    house_number = row.get("NUMERO CIVICO", "").strip()
    address_name = normalize_search_text(row.get("INDIRIZZO NORMALIZZATO", ""))
    if not address_name:
        address_name = normalize_search_text(row.get("CIVICO_NORM", ""))
    if road_only:
        address_name = normalize_search_text(row.get("CIVICO_NORM", "")) or address_name
    elif address_name and house_number and not address_name.endswith(f" {house_number}"):
        address_name = f"{address_name} {house_number}".strip()
    parts = [
        address_name,
        compact_spaces(city or row.get("COMUNE_NEW", "")),
        country,
    ]
    params = {
        "q": ", ".join(part for part in parts if part),
        "countrycodes": country_code,
        "format": "jsonv2",
        "limit": "1",
        "addressdetails": "1",
    }
    if email:
        params["email"] = email
    return params


def cache_key(params: dict[str, str]) -> str:
    return json.dumps(params, ensure_ascii=False, sort_keys=True)


def iter_geocoding_queries(
    row: dict[str, str],
    *,
    country: str,
    country_code: str,
    email: str | None,
) -> list[tuple[dict[str, str], str]]:
    primary_city = compact_spaces(row.get("COMUNE_NEW", ""))
    alternate_city = infer_alternate_city(row)
    queries: list[tuple[dict[str, str], str]] = [
        (
            build_structured_params(
                row,
                city=primary_city,
                country=country,
                country_code=country_code,
                email=email,
            ),
            "structured",
        ),
        (
            build_freeform_params(
                row,
                city=primary_city,
                country=country,
                country_code=country_code,
                email=email,
            ),
            "freeform",
        ),
        (
            build_freeform_params(
                row,
                city=primary_city,
                country=country,
                country_code=country_code,
                email=email,
                road_only=True,
            ),
            "road_only",
        ),
    ]

    if alternate_city:
        queries.extend(
            [
                (
                    build_structured_params(
                        row,
                        city=alternate_city,
                        country=country,
                        country_code=country_code,
                        email=email,
                    ),
                    "structured_alt_city",
                ),
                (
                    build_freeform_params(
                        row,
                        city=alternate_city,
                        country=country,
                        country_code=country_code,
                        email=email,
                    ),
                    "freeform_alt_city",
                ),
                (
                    build_freeform_params(
                        row,
                        city=alternate_city,
                        country=country,
                        country_code=country_code,
                        email=email,
                        road_only=True,
                    ),
                    "road_only_alt_city",
                ),
            ]
        )

    deduped: list[tuple[dict[str, str], str]] = []
    seen: set[str] = set()
    for params, query_type in queries:
        key = cache_key(params)
        if key in seen:
            continue
        seen.add(key)
        deduped.append((params, query_type))
    return deduped


def enforce_rate_limit(last_request_at: float | None, sleep_seconds: float) -> None:
    if last_request_at is None:
        return
    elapsed = time.monotonic() - last_request_at
    remaining = sleep_seconds - elapsed
    if remaining > 0:
        time.sleep(remaining)


def http_get_json(
    params: dict[str, str],
    *,
    user_agent: str,
    retries: int,
    sleep_seconds: float,
    last_request_at: float | None,
) -> tuple[list[dict[str, Any]], float]:
    for attempt in range(1, retries + 1):
        enforce_rate_limit(last_request_at, sleep_seconds)
        url = f"{NOMINATIM_URL}?{urlencode(params)}"
        request = Request(url, headers={"User-Agent": user_agent, "Accept": "application/json"})
        try:
            with urlopen(request, timeout=30) as response:
                payload = json.loads(response.read().decode("utf-8"))
                return payload, time.monotonic()
        except HTTPError as exc:
            retriable = exc.code in {429, 500, 502, 503, 504}
            if attempt == retries or not retriable:
                raise
            retry_after = exc.headers.get("Retry-After") if exc.headers else None
            wait_seconds = max(sleep_seconds, attempt * 2)
            if retry_after:
                try:
                    wait_seconds = max(wait_seconds, float(retry_after))
                except ValueError:
                    wait_seconds = max(wait_seconds, 15.0)
            elif exc.code == 429:
                wait_seconds = max(wait_seconds, attempt * 15.0)
            time.sleep(wait_seconds)
            last_request_at = time.monotonic()
        except URLError:
            if attempt == retries:
                raise
            time.sleep(max(sleep_seconds, attempt * 2))
            last_request_at = time.monotonic()
    raise RuntimeError("Tentativi esauriti senza risposta valida.")


def lookup_with_cache(
    params: dict[str, str],
    *,
    cache: dict[str, Any],
    user_agent: str,
    retries: int,
    sleep_seconds: float,
    last_request_at: float | None,
    dry_run: bool,
) -> tuple[list[dict[str, Any]], float | None, str]:
    key = cache_key(params)
    if key in cache:
        return cache[key], last_request_at, "cache"

    if dry_run:
        return [], last_request_at, "dry-run"

    results, last_request_at = http_get_json(
        params,
        user_agent=user_agent,
        retries=retries,
        sleep_seconds=sleep_seconds,
        last_request_at=last_request_at,
    )
    cache[key] = results
    return results, last_request_at, "api"


def enrich_row(
    row: dict[str, str],
    result: dict[str, Any] | None,
    *,
    status: str,
    source: str,
    query_type: str,
) -> dict[str, str]:
    enriched = dict(row)
    enriched["GEOCODER_STATUS"] = status
    enriched["GEOCODER_SOURCE"] = source
    enriched["GEOCODER_QUERY_TYPE"] = query_type
    strategy_map = {
        "structured": ("indirizzo_con_civico", "Ricerca standard con civico e comune normalizzato."),
        "freeform": ("indirizzo_libero_con_civico", "Ricerca libera con civico e comune normalizzato."),
        "road_only": ("solo_toponimo_senza_civico", "Coordinate cercate senza numero civico dopo fallimento della ricerca con civico."),
        "structured_alt_city": ("indirizzo_con_civico_comune_alternativo", "Ricerca con civico usando un comune dedotto dal testo grezzo."),
        "freeform_alt_city": ("indirizzo_libero_con_civico_comune_alternativo", "Ricerca libera con civico usando un comune dedotto dal testo grezzo."),
        "road_only_alt_city": ("solo_toponimo_senza_civico_comune_alternativo", "Coordinate cercate senza numero civico e con comune alternativo dedotto dal testo grezzo."),
    }
    strategy, note = strategy_map.get(
        query_type,
        ("strategia_non_specificata", "Strategia di geocoding non specificata."),
    )
    if status == "not_found" and query_type.startswith("road_only"):
        note = f"{note} Anche il fallback senza civico non ha prodotto risultati."
    enriched["GEOCODER_STRATEGY"] = strategy
    enriched["GEOCODER_NOTE"] = note

    if result is None:
        enriched["X"] = row.get("X", "")
        enriched["Y"] = row.get("Y", "")
        enriched["LAT"] = ""
        enriched["LON"] = ""
        enriched["DISPLAY_NAME"] = ""
        enriched["OSM_TYPE"] = ""
        enriched["OSM_ID"] = ""
        enriched["PLACE_CLASS"] = ""
        enriched["PLACE_TYPE"] = ""
        enriched["IMPORTANCE"] = ""
        return enriched

    lat = str(result.get("lat", ""))
    lon = str(result.get("lon", ""))
    enriched["X"] = lon
    enriched["Y"] = lat
    enriched["LAT"] = lat
    enriched["LON"] = lon
    enriched["DISPLAY_NAME"] = str(result.get("display_name", ""))
    enriched["OSM_TYPE"] = str(result.get("osm_type", ""))
    enriched["OSM_ID"] = str(result.get("osm_id", ""))
    enriched["PLACE_CLASS"] = str(result.get("class", ""))
    enriched["PLACE_TYPE"] = str(result.get("type", ""))
    enriched["IMPORTANCE"] = str(result.get("importance", ""))
    return enriched


def geocode_single_row(
    row: dict[str, str],
    *,
    cache: dict[str, Any],
    user_agent: str,
    retries: int,
    sleep_seconds: float,
    last_request_at: float | None,
    dry_run: bool,
    country: str,
    country_code: str,
    email: str | None,
) -> tuple[dict[str, str], float | None]:
    best_result = None
    source = "api"
    query_type = "structured"
    for params, candidate_query_type in iter_geocoding_queries(
        row,
        country=country,
        country_code=country_code,
        email=email,
    ):
        results, last_request_at, source = lookup_with_cache(
            params,
            cache=cache,
            user_agent=user_agent,
            retries=retries,
            sleep_seconds=sleep_seconds,
            last_request_at=last_request_at,
            dry_run=dry_run,
        )
        query_type = candidate_query_type
        best_result = results[0] if results else None
        if best_result is not None:
            break

    if best_result is None:
        output_row = enrich_row(
            row,
            None,
            status="not_found" if not dry_run else "dry_run",
            source=source,
            query_type=query_type,
        )
    else:
        output_row = enrich_row(
            row,
            best_result,
            status="matched",
            source=source,
            query_type=query_type,
        )

    return output_row, last_request_at


def write_generic_csv(path: Path, rows: list[dict[str, str]], base_headers: list[str]) -> None:
    extra_headers = [
        "GEOCODER_STATUS",
        "GEOCODER_SOURCE",
        "GEOCODER_QUERY_TYPE",
        "GEOCODER_STRATEGY",
        "GEOCODER_NOTE",
        "LAT",
        "LON",
        "DISPLAY_NAME",
        "OSM_TYPE",
        "OSM_ID",
        "PLACE_CLASS",
        "PLACE_TYPE",
        "IMPORTANCE",
    ]
    fieldnames: list[str] = []
    for header in base_headers + extra_headers:
        if header not in fieldnames:
            fieldnames.append(header)

    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def write_rows_with_headers(path: Path, rows: list[dict[str, str]], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def build_da_verificare_from_geocoded(validi_geocoded_csv: Path, da_verificare_csv: Path) -> dict[str, Any]:
    fieldnames, rows = read_csv_rows(validi_geocoded_csv)
    da_verificare_rows = [
        row for row in rows if (row.get("GEOCODER_STATUS") or "").strip() != "matched"
    ]
    write_rows_with_headers(da_verificare_csv, da_verificare_rows, fieldnames)
    return {
        "rows": len(da_verificare_rows),
        "fieldnames": fieldnames,
        "path": da_verificare_csv,
    }


def geocode_csv(
    input_csv: Path,
    output_csv: Path,
    *,
    cache_path: Path,
    email: str | None = None,
    user_agent: str = DEFAULT_USER_AGENT,
    country: str = "Italia",
    country_code: str = "it",
    sleep_seconds: float = 1.1,
    retries: int = 3,
    dry_run: bool = False,
    limit: int | None = None,
    progress_callback: Any | None = None,
) -> dict[str, Any]:
    headers, rows = read_csv_rows(input_csv)
    if limit is not None:
        rows = rows[:limit]

    cache = load_cache(cache_path)
    last_request_at: float | None = None
    output_rows: list[dict[str, str]] = []
    matched = 0
    not_found = 0

    for index, row in enumerate(rows, start=1):
        output_row, last_request_at = geocode_single_row(
            row,
            cache=cache,
            user_agent=user_agent,
            retries=retries,
            sleep_seconds=sleep_seconds,
            last_request_at=last_request_at,
            dry_run=dry_run,
            country=country,
            country_code=country_code,
            email=email,
        )
        if output_row["GEOCODER_STATUS"] == "matched":
            matched += 1
        else:
            not_found += 1

        output_rows.append(output_row)
        if progress_callback is not None:
            address = row.get("INDIRIZZO NORMALIZZATO", "").strip() or row.get("INDIRES", "").strip()
            progress_callback(index, len(rows), address, output_row["GEOCODER_STATUS"])

    write_generic_csv(output_csv, output_rows, headers)
    save_cache(cache_path, cache)
    return {
        "input_csv": input_csv,
        "output_csv": output_csv,
        "cache_path": cache_path,
        "rows": len(rows),
        "matched": matched,
        "not_found": not_found,
        "dry_run": dry_run,
    }


def geocode_csv_dedup_by_address(
    input_csv: Path,
    output_csv: Path,
    *,
    cache_path: Path,
    email: str | None = None,
    user_agent: str = DEFAULT_USER_AGENT,
    country: str = "Italia",
    country_code: str = "it",
    sleep_seconds: float = 1.1,
    retries: int = 3,
    dry_run: bool = False,
    progress_callback: Any | None = None,
) -> dict[str, Any]:
    headers, rows = read_csv_rows(input_csv)
    grouped_rows: dict[str, list[dict[str, str]]] = defaultdict(list)
    for row in rows:
        grouped_rows[normalized_address_key(row.get("INDIRIZZO NORMALIZZATO", ""))].append(row)

    cache = load_cache(cache_path)
    last_request_at: float | None = None
    geocoded_by_key: dict[str, dict[str, str]] = {}
    matched_groups = 0
    not_found_groups = 0

    for index, (key, group) in enumerate(grouped_rows.items(), start=1):
        representative = dict(group[0])
        geocoded_row, last_request_at = geocode_single_row(
            representative,
            cache=cache,
            user_agent=user_agent,
            retries=retries,
            sleep_seconds=sleep_seconds,
            last_request_at=last_request_at,
            dry_run=dry_run,
            country=country,
            country_code=country_code,
            email=email,
        )
        geocoded_row["GEOCODER_DUPLICATE_GROUP_SIZE"] = str(len(group))
        geocoded_by_key[key] = geocoded_row
        if geocoded_row["GEOCODER_STATUS"] == "matched":
            matched_groups += 1
        else:
            not_found_groups += 1
        if progress_callback is not None:
            address = representative.get("INDIRIZZO NORMALIZZATO", "").strip() or representative.get("INDIRES", "").strip()
            progress_callback(index, len(grouped_rows), address, geocoded_row["GEOCODER_STATUS"])

    output_rows: list[dict[str, str]] = []
    for row in rows:
        key = normalized_address_key(row.get("INDIRIZZO NORMALIZZATO", ""))
        geocoded = dict(geocoded_by_key[key])
        geocoded["_row"] = row.get("_row", geocoded.get("_row", ""))
        geocoded["_category"] = row.get("_category", geocoded.get("_category", ""))
        geocoded["_issues"] = row.get("_issues", geocoded.get("_issues", ""))
        geocoded["INDIRES"] = row.get("INDIRES", geocoded.get("INDIRES", ""))
        geocoded["TYPE"] = row.get("TYPE", geocoded.get("TYPE", ""))
        geocoded["CIVICO_NORM"] = row.get("CIVICO_NORM", geocoded.get("CIVICO_NORM", ""))
        geocoded["NUMERO CIVICO"] = row.get("NUMERO CIVICO", geocoded.get("NUMERO CIVICO", ""))
        geocoded["INDIRIZZO NORMALIZZATO"] = row.get("INDIRIZZO NORMALIZZATO", geocoded.get("INDIRIZZO NORMALIZZATO", ""))
        geocoded["COMUNE_NEW"] = row.get("COMUNE_NEW", geocoded.get("COMUNE_NEW", ""))
        geocoded["GEOCODER_REUSED_FOR_DUPLICATE"] = "yes" if int(geocoded["GEOCODER_DUPLICATE_GROUP_SIZE"]) > 1 else "no"
        output_rows.append(geocoded)

    save_cache(cache_path, cache)
    extra_headers = headers + ["GEOCODER_DUPLICATE_GROUP_SIZE", "GEOCODER_REUSED_FOR_DUPLICATE"]
    write_generic_csv(output_csv, output_rows, extra_headers)
    return {
        "input_csv": input_csv,
        "output_csv": output_csv,
        "cache_path": cache_path,
        "rows": len(rows),
        "unique_addresses": len(grouped_rows),
        "matched_groups": matched_groups,
        "not_found_groups": not_found_groups,
        "dry_run": dry_run,
    }


def run_full_pipeline(
    input_file: Path,
    output_dir: Path,
    *,
    sheet_name: str | None = None,
    geocoder_email: str | None = None,
    user_agent: str = DEFAULT_USER_AGENT,
    country: str = "Italia",
    country_code: str = "it",
    sleep_seconds: float = 1.1,
    retries: int = 3,
    dry_run: bool = False,
) -> dict[str, Any]:
    analysis = analyze_input_file(input_file, output_dir, sheet_name=sheet_name)
    stem = input_file.stem
    geocoded_csv = output_dir / f"{stem}_validi_geocoded.csv"
    cache_path = output_dir / "nominatim_cache.json"

    geocode = geocode_csv_dedup_by_address(
        analysis["paths"]["validi"],
        geocoded_csv,
        cache_path=cache_path,
        email=geocoder_email,
        user_agent=user_agent,
        country=country,
        country_code=country_code,
        sleep_seconds=sleep_seconds,
        retries=retries,
        dry_run=dry_run,
    )
    da_verificare = build_da_verificare_from_geocoded(
        geocoded_csv,
        analysis["paths"]["da_verificare"],
    )

    analysis["paths"]["validi_geocoded"] = geocoded_csv
    analysis["paths"]["cache"] = cache_path
    analysis["geocode"] = geocode
    analysis["da_verificare"] = da_verificare
    return analysis


def copy_input_to_output(input_file: Path, output_dir: Path) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    destination = output_dir / input_file.name
    shutil.copy2(input_file, destination)
    return destination
